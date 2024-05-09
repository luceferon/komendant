import configparser
import locale
from datetime import datetime
import openpyxl
import pymysql
import win32com.client
from PyQt5.QtWidgets import (QApplication, QHeaderView, QLabel, QMessageBox)
from PyQt5.uic import loadUiType
from PyQt5.QtGui import QColor, QStandardItemModel, QStandardItem, QFont
from PyQt5.QtCore import Qt
import sys
import os
import calendar
from openpyxl.reader.excel import load_workbook

from settings import load_settings

Ui_MainWindow, QMainWindow = loadUiType("ZPui.ui")


class ExcelPrinter:
    def __init__(self, server_address, smb_share):
        config = configparser.ConfigParser()
        config.read('conf.ini')
        self.server_address = config.get('FileServ', 'server')
        self.server_address = server_address
        self.smb_share = '\\smb_share\\komendant\\'

    def print_excel_file(self):
        file_path = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'zptest.xlsx')
        workbook = load_workbook(file_path)
        worksheet = workbook.active or workbook['Sheet1']
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(file_path)
        workbook.PrintOut()
        workbook.Close()
        excel.Quit()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.PBPrint.clicked.connect(self.transferRowsToTBTemp)
        self.PBAdd.clicked.connect(self.addNewRow)
        self.PBUncheck.clicked.connect(self.uncheckAllItems)
        config = configparser.ConfigParser()
        config.read('conf.ini')
        self.server_address = config.get('FileServ', 'server')
        self.smb_share = '\\smb_share\\komendant\\'
        config_file_path = f'\\\\{self.server_address}\\smb_share\\komendant\\conf.ini'
        settings = load_settings(config_file_path)
        if settings is not None:
            self.server, self.user, self.password, self.db, self.NachUch, self.Uchastok, self.Kladovchik, self.SysAdmin = settings
            connection = pymysql.connect(host=self.server, port=3306, user=self.user, password=self.password,
                                         db=self.db)
        else:
            QMessageBox.warning(self, "Warning", f"Файл настроек {config_file_path} не найден.", QMessageBox.Ok)
            sys.exit()

        self.cursor = connection.cursor()
        self.cursor.execute('SELECT * FROM balki')
        self.data_balki = self.cursor.fetchall()
        self.cursor.execute('SELECT * FROM obchaga')
        self.data_obchaga = self.cursor.fetchall()
        self.model_all = QStandardItemModel()
        self.model_all.itemChanged.connect(self.updateCheckedCount)

        for row_data in self.data_balki + self.data_obchaga:
            row = []
            self.check_item = QStandardItem()
            self.check_item.setCheckable(True)
            self.check_item.setEditable(False)
            row.append(self.check_item)
            for item in row_data:
                standard_item = QStandardItem(str(item))
                row.append(standard_item)
            percent_item = QStandardItem("100%")
            row.append(percent_item)
            self.model_all.appendRow(row)

        self.TVGlav.setModel(self.model_all)
        self.model_all.removeColumn(8)
        self.model_all.removeColumn(7)
        self.model_all.removeColumn(5)
        self.model_all.removeColumn(4)
        self.model_all.removeColumn(2)
        self.model_all.removeColumn(1)
        self.model_all.setHorizontalHeaderLabels(
            ["Включить", "ФИО", "Должность", "Сумма"])
        self.TVGlav.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        new_row = [
            QStandardItem(self.check_item),
            QStandardItem(self.NachUch),
            QStandardItem("Начальник участка"),
            QStandardItem("П/О 50 т.р.")
        ]
        self.model_all.appendRow(new_row)
        self.cursor.close()
        connection.close()
        self.LESearch.textChanged.connect(self.searchTable)
        self.sortTableByFIO()
        self.model_TBTemp = QStandardItemModel()
        self.TBTemp.setModel(self.model_TBTemp)
        self.model_TBTemp.setHorizontalHeaderLabels(["ФИО", "Должность", "Сумма"])
        self.TBTemp.setVisible(False)
        locale.setlocale(locale.LC_ALL, 'ru_RU.utf8')

    def addNewRow(self):
        # Создание новой пустой строки
        new_row = [
            QStandardItem(self.check_item),
            QStandardItem(""),
            QStandardItem(""),
            QStandardItem("100%")
        ]
        self.model_all.appendRow(new_row)

    def uncheckAllItems(self):
        # Кнопка снять выделение
        for row in range(self.model_all.rowCount()):
            item = self.model_all.item(row, 0)
            if item and item.isCheckable():
                item.setCheckState(Qt.Unchecked)

    def transferRowsToTBTemp(self):
        row_count = self.model_TBTemp.rowCount()
        for row_index in range(self.model_all.rowCount()):
            check_item = self.model_all.item(row_index, 0)
            if check_item.checkState() == Qt.Checked:
                fio_item = self.model_all.item(row_index, 1)
                position_item = self.model_all.item(row_index, 2)
                sum_item = self.model_all.item(row_index, 3)
                if fio_item.text() and position_item.text() and sum_item.text():
                    row_data = [
                        QStandardItem(str(row_count + 1)),
                        QStandardItem(fio_item.text()),
                        QStandardItem(position_item.text()),
                        QStandardItem(sum_item.text()
                                      )
                    ]
                    self.model_TBTemp.appendRow(row_data)
                    row_count += 1
                else:
                    print(f"Одна из ячеек в строке {row_index} пуста. Строка не добавлена.")
        self.print_and_save_excel()

    def sortTableByFIO(self):
        # Сортировка таблицы по второму столбцу (ФИО)
        self.model_all.sort(1, Qt.AscendingOrder)

    def searchTable(self, searchText):
        # Поиск по таблице
        currentIndex = self.TVGlav.currentIndex()
        rowCount = self.TVGlav.model().rowCount()
        if rowCount > 0:
            for row in range(rowCount):
                self.TVGlav.model().setData(self.TVGlav.model().index(row, 1), QColor("white"), Qt.BackgroundRole)
            for row in range(rowCount):
                item = self.TVGlav.model().item(row, 1)
                if item is not None:
                    if searchText.lower() in item.text().lower():
                        self.TVGlav.model().setData(self.TVGlav.model().index(row, 1), QColor("yellow"),
                                                    Qt.BackgroundRole)
            if rowCount > 0:
                for row in range(rowCount):
                    if self.TVGlav.model().item(row, 1).text().lower().find(searchText.lower()) != -1:
                        self.TVGlav.setCurrentIndex(self.TVGlav.model().index(row, 1))
                        break
            if currentIndex.isValid():
                self.TVGlav.setCurrentIndex(currentIndex)
        else:
            print("Модель данных пуста.")
        if searchText == "":
            for row in range(rowCount):
                self.TVGlav.model().setData(self.TVGlav.model().index(row, 1), QColor("white"),
                                            Qt.BackgroundRole)

    def updateCheckedCount(self, item):
        if item.column() == 0:
            checked_count = 0
            for row in range(self.model_all.rowCount()):
                item = self.model_all.item(row, 0)
                if item is not None and item.checkState() == Qt.Checked:
                    checked_count += 1
            self.LCount.setText(" {}".format(checked_count))
            if checked_count >= 43:
                self.LCount.setFont(QFont("Arial", 16, QFont.Bold | QFont.StyleItalic))
                self.LCount.setStyleSheet("color: red;")
                self.PBPrint.setEnabled(False)
            else:
                self.LCount.setFont(QFont("Arial", 12, QFont.Bold | QFont.StyleItalic))
                self.LCount.setStyleSheet("color: black;")
                self.PBPrint.setEnabled(True)

    def print_and_save_excel(self):
        # Печатаем списки
        file_path = os.path.join(f'\\\\{self.server_address}\\smb_share\\komendant\\zptest.xlsx')
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=5, max_row=46, min_col=1, max_col=sheet.max_column):
            for cell in row:
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        sheet.unmerge_cells(merged_range.coord)
                        break
                if cell.value is not None or cell.has_style:
                    cell.value = None
        current_year = datetime.now().strftime("%Y")
        prev_month = calendar.month_name[datetime.now().month - 1] if datetime.now().month != 1 else \
            calendar.month_name[12]
        prev_year = current_year if datetime.now().month == 1 else current_year
        sheet['A1'] = f"Список работников на з\плату за {prev_month} {prev_year} г."
        sheet['C3'] = self.uchastok
        sheet['C48'] = self.NachUch
        for row_index in range(5, self.model_TBTemp.rowCount() + 5):
            for column_index in range(1, self.model_TBTemp.columnCount() + 1):
                item = self.model_TBTemp.item(row_index - 5, column_index - 1)
                if item:
                    sheet.cell(row=row_index, column=column_index, value=item.text())
        workbook.save(file_path)
        printer = ExcelPrinter(self.server_address, self.smb_share)
        printer.print_excel_file()
        self.model_TBTemp.clear()

    def closeEvent(self, event):
        self.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
