import configparser
import subprocess
import openpyxl
import pymysql
import win32com.client
import xlsxwriter
import re
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QHeaderView, QCheckBox, QDialog, QVBoxLayout, QDialogButtonBox,
                             QMessageBox, QCalendarWidget, QPushButton, QLineEdit)
from PyQt5.uic import loadUiType
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QColor, QRegExpValidator
from PyQt5.QtCore import Qt, QDate, QRegExp
import sys
import os
from settings import load_settings

from openpyxl.reader.excel import load_workbook

Ui_MainWindow, QMainWindow = loadUiType("Zaselenieui.ui")
Ui_Dialog, QDialog = loadUiType("OtpuskUI.ui")


class ExcelPrinter:
    def __init__(self, server_address, smb_share):
        config = configparser.ConfigParser()
        config.read('conf.ini')
        self.server_address = config.get('FileServ', 'server')
        self.smb_share = '\\smb_share\\komendant\\'

    def print_excel_file(self, file_path):
        workbook = load_workbook(file_path)
        worksheet = workbook.active or workbook['Sheet1']
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(file_path)
        workbook.PrintOut()
        workbook.Close()
        excel.Quit()


# Диалог ввода номера телефона
class PhoneNumberInputDialog(QDialog):
    def __init__(self, initial_phone=None):
        super().__init__()
        self.setWindowTitle("Телефон")
        layout = QVBoxLayout()
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Введите номер телефона в формате 8(***)***-**-**")
        self.phone_input.setValidator(
            QRegExpValidator(QRegExp(r'^(\d{1})[\s-]?(\d{3})[\s-]?(\d{3})[\s-]?(\d{2})[\s-]?(\d{2})$')))
        self.phone_input.textChanged.connect(self.validate_phone)
        layout.addWidget(self.phone_input)
        self.ok_button = QPushButton('OK')
        self.ok_button.clicked.connect(self.accept)
        layout.addWidget(self.ok_button)
        self.setLayout(layout)
        if initial_phone:
            self.phone_input.setText(initial_phone)

    def validate_phone(self):
        text = self.phone_input.text()
        phone_digits = ''.join(filter(str.isdigit, text))
        if len(phone_digits) == 11:
            formatted_phone = '8({}){}-{}-{}'.format(phone_digits[1:4], phone_digits[4:7], phone_digits[7:9],
                                                     phone_digits[9:11])
            self.phone_input.setText(formatted_phone)
            self.ok_button.setEnabled(True)
        else:
            self.ok_button.setEnabled(False)


# Диалог ввода даты
class DateInputDialog(QDialog):
    def __init__(self, initial_date=None):
        super().__init__()
        self.setWindowTitle("Дата")
        layout = QVBoxLayout()
        self.cal = QCalendarWidget()
        layout.addWidget(self.cal)
        self.ok_button = QPushButton('OK')
        self.ok_button.clicked.connect(self.accept)
        layout.addWidget(self.ok_button)
        self.setLayout(layout)
        if initial_date:
            self.cal.setSelectedDate(initial_date)

    def dateValue(self):
        return self.cal.selectedDate()


# Диалог экспорта
class ExportDialog(QDialog):
    def __init__(self, columns):
        super().__init__()
        self.setWindowTitle("Выберите столбцы для экспорта")
        self.layout = QVBoxLayout()
        for i, column in enumerate(columns):
            checkbox = QCheckBox(column)
            self.layout.addWidget(checkbox)
            setattr(self, f"cb_{i}", checkbox)
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        self.layout.addWidget(self.buttonBox)
        self.setLayout(self.layout)


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.Company = None
        self.Dolznost = None
        self.FIO = None
        self.smb_share = '\\smb_share\\komendant\\'
        self.setupUi(self)
        self.state_obhod = 'Отпуск'
        self.PBclose.clicked.connect(self.close)
        self.PBZP.clicked.connect(self.zpwindow)
        self.PBOtpusk.clicked.connect(self.otpusk_dialog)
        self.TVBalki.doubleClicked.connect(self.phone_double_clicked)
        self.TVObchaga.doubleClicked.connect(self.phone_double_clicked)
        self.TVBalki.doubleClicked.connect(self.edit_date_balki)
        self.TVObchaga.doubleClicked.connect(self.edit_date_obchaga)

        config = configparser.ConfigParser()
        config.read('conf.ini')
        self.server_address = config.get('FileServ', 'server')
        config_file_path = f'\\\\{self.server_address}\\smb_share\\komendant\\conf.ini'
        settings = load_settings(config_file_path)

        if settings is not None:
            self.server, self.user, self.password, self.db, self.NachUch, self.Uchastok, self.Kladovchik, self.SysAdmin = settings
            connection = pymysql.connect(host=self.server, port=3306, user=self.user, password=self.password, db=self.db)
        else:
            QMessageBox.warning(self, "Warning", f"Файл настроек {config_file_path} не найден.", QMessageBox.Ok)
            sys.exit()

        self.cursor = connection.cursor()
        self.cursor.execute('SELECT * FROM balki')
        self.data = self.cursor.fetchall()
        self.cursor.execute('SELECT * FROM obchaga')
        self.data_obchaga = self.cursor.fetchall()

        model = QStandardItemModel()
        model_obchaga = QStandardItemModel()

        self.headers = ['Номер', 'Кол-во_мест', 'ФИО_сотрудника', 'Комментарий', 'Дата_заезда_отпуска', 'Должность',
                        'Организация', 'Контактный_телефон', ]
        for header in self.headers:
            item = QStandardItem(header)
            item.setEditable(False)
            model.setHorizontalHeaderItem(self.headers.index(header), item)
        self.headers_obchaga = ['Номер', 'Кол-во_мест', 'ФИО_сотрудника', 'Комментарий', 'Дата_заезда_отпуска',
                                'Должность',
                                'Организация', 'Контактный_телефон', ]
        for header in self.headers_obchaga:
            item = QStandardItem(header)
            item.setEditable(False)
            model_obchaga.setHorizontalHeaderItem(self.headers_obchaga.index(header), item)

        def custom_sort_key(x):
            value = x[0]
            try:
                return (int(value), "")
            except ValueError:
                return (
                    float("inf"), value)

        sorted_data = sorted(self.data, key=lambda x: custom_sort_key(x))
        for row_data in sorted_data:
            row = []
            for column_index, column_data in enumerate(row_data):
                item = QStandardItem(str(column_data))
                if column_index in [1, 3, 5, 6]:
                    item.setEditable(True)
                else:
                    item.setEditable(False)
                row.append(item)
            model.appendRow(row)
        sorted_data_obchaga = sorted(self.data_obchaga, key=lambda x: custom_sort_key(x))
        for row_data in sorted_data_obchaga:
            row = []
            for column_index, column_data in enumerate(row_data):
                item = QStandardItem(str(column_data))
                if column_index in [1, 3, 5, 6]:
                    item.setEditable(True)
                else:
                    item.setEditable(False)
                row.append(item)
            model_obchaga.appendRow(row)

        self.TVBalki.setModel(model)
        self.TVObchaga.setModel(model_obchaga)
        self.TVBalki.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.TVObchaga.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.PBExport.clicked.connect(self.export_data)
        self.cursor.close()
        connection.close()

        self.LESearch.textChanged.connect(self.searchTable)

        def search_and_highlight():
            # поиск и подсветка отпускников и опасдунов
            search_text = "отпуск"
            today = datetime.now().date()
            date_regex = r"(\d{2}\.\d{2}\.\d{4}) - (\d{2}\.\d{2}\.\d{4})|(\d{2}\.\d{2}\.\d{4})"
            for row in range(self.TVBalki.model().rowCount()):
                item = self.TVBalki.model().item(row, 3)
                if item and re.search(search_text, item.text().lower()):
                    item.setBackground(QColor('blue'))
                date_item = self.TVBalki.model().item(row, 4)
                if date_item:
                    match = re.match(date_regex, date_item.text())
                    if match:
                        if match.group(3):  # Дата в формате "дд.мм.гггг"
                            pass
                        else:  # Дата в формате "дд.мм.гггг - дд.мм.гггг"
                            start_date_str = match.group(1)  # Начальная дата
                            end_date_str = match.group(2)  # Конечная дата
                            if start_date_str:
                                start_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
                            else:
                                start_date = None
                            end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
                            if end_date < today:
                                date_item.setBackground(QColor('red'))
            for row in range(self.TVObchaga.model().rowCount()):
                item = self.TVObchaga.model().item(row, 3)
                if item and re.search(search_text, item.text().lower()):
                    item.setBackground(QColor('blue'))  # Цвет заливки Отпускников
                date_item = self.TVObchaga.model().item(row, 4)
                if date_item:
                    match = re.match(date_regex, date_item.text())
                    if match:
                        if match.group(3):  # Дата в формате "дд.мм.гггг"
                            pass
                        else:  # Дата в формате "дд.мм.гггг - дд.мм.гггг"
                            start_date_str = match.group(1)
                            end_date_str = match.group(2)
                            if start_date_str:
                                start_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
                            else:
                                start_date = None
                            end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
                            if end_date < today:
                                date_item.setBackground(QColor('red'))  # Цвет заливки опаздунов

        search_and_highlight()

    # Редактируем телефон
    def phone_double_clicked(self, index):
        if index.column() == 7:
            phone = index.model().data(index, Qt.DisplayRole)
            dialog = PhoneNumberInputDialog(phone)
            if dialog.exec_():
                new_phone = dialog.phone_input.text()
                index.model().setData(index, new_phone, Qt.DisplayRole)

    # Редактируем дату
    def edit_date_balki(self, index):
        current_date_str = index.model().data(index, Qt.DisplayRole)
        if not current_date_str:
            return
        dates = current_date_str.split(' - ')  # Проверяем формат "дд.мм.гггг - дд.мм.гггг"
        if len(dates) == 2:
            current_date_str = dates[0]
        try:
            current_date = datetime.strptime(current_date_str, '%d.%m.%Y').date()
        except ValueError:
            return
        dialog = DateInputDialog(QDate.fromString(current_date_str, 'dd.MM.yyyy'))
        if dialog.exec_() == QDialog.Accepted:
            new_date = dialog.dateValue().toString('dd.MM.yyyy')
            if len(dates) == 2:
                new_date_str = f'{new_date}'
            else:
                new_date_str = new_date
            index.model().setData(index, new_date_str, Qt.DisplayRole)

    # Редактируем дату
    def edit_date_obchaga(self, index):
        current_date_str = index.model().data(index, Qt.DisplayRole)
        if not current_date_str:
            return
        dates = current_date_str.split(' - ')  # Проверяем формат "дд.мм.гггг - дд.мм.гггг"
        if len(dates) == 2:
            current_date_str = dates[0]
        try:
            current_date = datetime.strptime(current_date_str, '%d.%m.%Y').date()
        except ValueError:
            return
        dialog = DateInputDialog(QDate.fromString(current_date_str, 'dd.MM.yyyy'))
        if dialog.exec_() == QDialog.Accepted:
            new_date = dialog.dateValue().toString('dd.MM.yyyy')
            if len(dates) == 2:
                new_date_str = f'{new_date}'
            else:
                new_date_str = new_date
            index.model().setData(index, new_date_str, Qt.DisplayRole)

    # Экспорт в эксель списка сотрудников
    def export_data(self):
        columns = ['Номер', 'ФИО_сотрудника', 'Комментарий', 'Дата_заезда_отпуска', 'Должность',
                   'Организация', 'Контактный_телефон', ]
        export_dialog = ExportDialog(columns)
        result = export_dialog.exec_()
        if result == QDialog.Accepted:
            selected_columns = []
            for i, column in enumerate(columns):
                if getattr(export_dialog, f"cb_{i}").isChecked():
                    selected_columns.append(column)
            workbook = xlsxwriter.Workbook('C:\Комендант\Списки сотрудников.xlsx')
            worksheet = workbook.add_worksheet()
            bold = workbook.add_format({'bold': True})
            header_format = workbook.add_format(
                {'font_name': 'Arial', 'font_size': 12, 'text_wrap': False, 'align': 'center'})
            for col, column in enumerate(selected_columns, start=1):
                worksheet.write(0, col, column, header_format)
            row = 1
            for data_row in self.data:
                col = 1
                for column in selected_columns:
                    col_index = self.headers.index(column)
                    data = data_row[col_index]
                    if isinstance(data, str) and len(data) > 50:
                        worksheet.set_column(col, col, 30)
                    elif isinstance(data, int) or isinstance(data, float):
                        worksheet.set_column(col, col, 10)
                    worksheet.write(row, col, data_row[col_index])
                    col += 1
                row += 1
            for data_row in self.data_obchaga:
                col = 1
                for column in selected_columns:
                    col_index = self.headers_obchaga.index(column)
                    data = data_row[col_index]
                    if isinstance(data, str) and len(data) > 50:
                        worksheet.set_column(col, col, 30)
                    elif isinstance(data, int) or isinstance(data, float):
                        worksheet.set_column(col, col, 10)
                    worksheet.write(row, col, data_row[col_index])
                    col += 1
                row += 1
            workbook.close()
            QMessageBox.information(self, "Экспорт завершен",
                                    "Данные успешно экспортированы в файл 'C:\Комендант\Списки сотрудников.xlsx'.")

    def searchTable(self, searchText):
        # Поиск по таблицам
        currentIndex = self.TVBalki.currentIndex()
        rowCount = self.TVBalki.model().rowCount()
        if rowCount > 0:
            for row in range(rowCount):
                self.TVBalki.model().setData(self.TVBalki.model().index(row, 2), QColor("white"), Qt.BackgroundRole)
            for row in range(rowCount):
                item = self.TVBalki.model().item(row, 2)
                if item is not None:
                    if searchText.lower() in item.text().lower():
                        self.TVBalki.model().setData(self.TVBalki.model().index(row, 2), QColor("yellow"),
                                                     Qt.BackgroundRole)
            if rowCount > 0:
                for row in range(rowCount):
                    if self.TVBalki.model().item(row, 2).text().lower().find(searchText.lower()) != -1:
                        self.TVBalki.setCurrentIndex(self.TVBalki.model().index(row, 0))
                        break
            rowCount_obchaga = self.TVObchaga.model().rowCount()
            if rowCount_obchaga > 0:
                for row in range(rowCount_obchaga):
                    self.TVObchaga.model().setData(self.TVObchaga.model().index(row, 2), QColor("white"),
                                                   Qt.BackgroundRole)
                for row in range(rowCount_obchaga):
                    item = self.TVObchaga.model().item(row, 2)
                    if item is not None:
                        if searchText.lower() in item.text().lower():
                            self.TVObchaga.model().setData(self.TVObchaga.model().index(row, 2), QColor("yellow"),
                                                           Qt.BackgroundRole)
                if rowCount_obchaga > 0:
                    for row in range(rowCount_obchaga):
                        if self.TVObchaga.model().item(row, 2).text().lower().find(searchText.lower()) != -1:
                            self.TVObchaga.setCurrentIndex(self.TVObchaga.model().index(row, 0))
                            break
            if currentIndex.isValid():
                self.TVBalki.setCurrentIndex(currentIndex)
        else:
            print("Модель данных пуста.")
        if searchText == "":
            for row in range(rowCount):
                self.TVBalki.model().setData(self.TVBalki.model().index(row, 2), QColor("white"),
                                             Qt.BackgroundRole)

    def save_data_to_db(self):
        connection = pymysql.connect(host=self.server, port=3306, user=self.user, password=self.password, db=self.db)
        self.cursor = connection.cursor()
        self.cursor.execute('DELETE FROM balki')
        model = self.TVBalki.model()
        rows = model.rowCount()
        for row in range(rows):
            data = []
            for column in range(model.columnCount()):
                item = model.item(row, column)
                data.append(item.text())
            self.cursor.execute('INSERT INTO balki VALUES (%s, %s, %s, %s, %s, %s, %s, %s)', data)
        self.cursor.execute('DELETE FROM obchaga')
        model_obchaga = self.TVObchaga.model()
        rows_obchaga = model_obchaga.rowCount()
        for row in range(rows_obchaga):
            data = []
            for column in range(model_obchaga.columnCount()):
                item = model_obchaga.item(row, column)
                data.append(item.text())
            self.cursor.execute('INSERT INTO obchaga VALUES (%s, %s, %s, %s, %s, %s, %s, %s)', data)
        connection.commit()
        self.cursor.close()
        connection.close()

    def zpwindow(self, event):
        self.save_data_to_db()
        subprocess.Popen(["python", "ZP.py"])

    def otpusk_dialog(self):
        dialog = QDialog()
        ui = Ui_Dialog()
        ui.setupUi(dialog)
        current_date = datetime.now().date()
        ui.DENachalo.setDate(current_date)
        ui.DEKonec.setDate(current_date + timedelta(days=30))  # Добавляем месяц (30 дней)

        # прячем т.к. увольняем
        def handle_uval_state_changed(state):
            if state:
                ui.DENachalo.hide()
                ui.DEKonec.hide()
                ui.CBObhod.hide()
                ui.CBDengi.hide()
                ui.DBOtpusk.hide()
                ui.label_2.hide()
                ui.label_3.hide()
                ui.label_4.hide()
                self.state_obhod = 'Увольнение'
            else:
                ui.DENachalo.show()
                ui.DEKonec.show()
                ui.CBObhod.show()
                ui.CBDengi.show()
                ui.DBOtpusk.show()
                ui.label_2.show()
                ui.label_3.show()
                ui.label_4.show()
                self.state_obhod = 'Отпуск'

        ui.CBUval.stateChanged.connect(handle_uval_state_changed)

        selected_index = None
        if self.tabWidget.currentIndex() == 0:  # вкладка tabBalki
            if self.TVBalki.selectedIndexes():  # Проверка наличия выделенных ячеек
                selected_index = self.TVBalki.selectedIndexes()[0]
        elif self.tabWidget.currentIndex() == 1:  # вкладка tabObchaga
            if self.TVObchaga.selectedIndexes():  # Проверка наличия выделенных ячеек
                selected_index = self.TVObchaga.selectedIndexes()[0]
        if selected_index is not None:
            if selected_index.column() == 2:  # проверка что выбрана ячейка в столбце ФИО
                self.FIO = selected_index.data()
                self.Dolznost = selected_index.sibling(selected_index.row(), 5).data()
                self.Company = selected_index.sibling(selected_index.row(), 6).data()

                def handle_button_click():
                    self.Dni = ui.SBDni.value()
                    self.Uval = ui.CBUval.isChecked()
                    self.Nachalo = ui.DENachalo.date()
                    self.Konec = ui.DEKonec.date()
                    self.Obhod = ui.CBObhod.isChecked()
                    self.Dengi = ui.CBDengi.isChecked()
                    self.Otpusk = ui.DBOtpusk.isChecked()

                    model = selected_index.model()
                    model.setData(model.index(selected_index.row(), 3), "Отпуск")
                    model.setData(model.index(selected_index.row(), 4),
                                  f"{self.Nachalo.toString('dd.MM.yyyy')} - {self.Konec.toString('dd.MM.yyyy')}")

                    # печатаем заявление на деньги
                    if ui.buttonBox.accepted.connect(
                            handle_button_click) and self.Dengi:  # по нажатию ОК проверяем стоит ли галка на CBDengi
                        zpotpusk_doc = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'zpotpusk.xlsx')
                        self.print_doc_zpotpusk(zpotpusk_doc)

                    # печатаем заявление на обходной
                    if ui.buttonBox.accepted.connect(
                            handle_button_click) and self.Obhod:  # по нажатию ОК проверяем стоит ли галка на CBObhod
                        obhod_doc = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'obhod.xlsx')
                        self.print_doc_obhod(obhod_doc)

                    # печатаем заявление на отпуск
                    if ui.buttonBox.accepted.connect(
                            handle_button_click) and self.Otpusk:  # по нажатию ОК проверяем стоит ли галка на CBOtpusk
                        otpusk_doc = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'otpusk.xlsx')
                        self.print_doc_otpusk(otpusk_doc)

                    # печатаем заявление на увал
                    if ui.buttonBox.accepted.connect(
                            handle_button_click) and self.Uval:  # по нажатию ОК проверяем стоит ли галка на CBUval
                        uval_doc = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'uval.xlsx')
                        self.print_doc_uval(uval_doc)
                        obhod_doc = os.path.join(f'\\\\{self.server_address}\\{self.smb_share}', 'obhod.xlsx')
                        self.print_doc_obhod(obhod_doc)

                ui.buttonBox.accepted.connect(handle_button_click)
                dialog.exec_()
            else:
                QMessageBox.warning(self, "Warning", "Необходимо выбрать сотрудника", QMessageBox.Ok)
        else:
            QMessageBox.warning(self, "Warning", "Необходимо выбрать сотрудника", QMessageBox.Ok)
        return self.FIO, self.Dolznost, self.Company

    def print_doc_zpotpusk(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Формирование статичных данных
        sheet['A5'] = f"От {self.Dolznost}"
        sheet['A6'] = self.FIO

        # Проверка условия для Company
        if self.Company == 'ООО "А-Сервис"':
            sheet['A3'] = 'ООО "А-Сервис"'
            sheet['A4'] = "Белову В. Ю."

        if self.Company == 'ООО "Еда"':
            sheet['A3'] = 'ООО "Еда"'
            sheet['A4'] = "Потапову М. Г."

        if self.Company == 'ООО "Буфет"':
            sheet['A3'] = 'ООО "Буфет"'
            sheet['A4'] = "Потапову Г. В."

        if self.Company == 'ООО "ГМК Ангара"':
            sheet['A3'] = 'ООО "ГМК Ангара"'
            sheet['A4'] = "Калдарумиди Р. П."

        if self.Company == 'ООО "КомплектСервис"':
            sheet['A3'] = 'ООО "КомплектСервис"'
            sheet['A4'] = "Потриденный В. Ф."

        if self.Company == 'ООО "МоторСервис"':
            sheet['A3'] = 'ООО "МоторСервис"'
            sheet['A4'] = "Борисюк К. М."

        if self.Company == 'ООО "Сисим"':
            sheet['A3'] = 'ООО "Сисим"'
            sheet['A4'] = "Ковалькову М. Н."

        if self.Company == 'ООО "СпецПодряд"':
            sheet['A3'] = 'ООО "СпецПодряд"'
            sheet['A4'] = "Осс А. В."

        if self.Company == 'ООО "КрасИнтегра"':
            sheet['A3'] = 'ООО "КрасИнтегра"'
            sheet['A4'] = "Сапину В. Д."

        # Сохранение файла
        workbook.save(file_path)

        # Печать файла
        printer = ExcelPrinter(self.server_address, self.smb_share)
        printer.print_excel_file(file_path)

    def print_doc_otpusk(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Формирование статичных данных
        sheet['A5'] = f"От {self.Dolznost}"
        sheet['A6'] = self.FIO
        sheet['A15'] = f"С {self.Nachalo.toString('dd.MM.yyyy')}"
        sheet['D15'] = f"По  {self.Konec.toString('dd.MM.yyyy')}"

        # Проверка условия для Company
        if self.Company == 'ООО "А-Сервис"':
            sheet['A3'] = 'ООО "А-Сервис"'
            sheet['A4'] = "Белову В. Ю."

        if self.Company == 'ООО "Еда"':
            sheet['A3'] = 'ООО "Еда"'
            sheet['A4'] = "Потапову М. Г."

        if self.Company == 'ООО "Буфет"':
            sheet['A3'] = 'ООО "Буфет"'
            sheet['A4'] = "Потапову Г. В."

        if self.Company == 'ООО "ГМК Ангара"':
            sheet['A3'] = 'ООО "ГМК Ангара"'
            sheet['A4'] = "Калдарумиди Р. П."

        if self.Company == 'ООО "КомплектСервис"':
            sheet['A3'] = 'ООО "КомплектСервис"'
            sheet['A4'] = "Потриденный В. Ф."

        if self.Company == 'ООО "МоторСервис"':
            sheet['A3'] = 'ООО "МоторСервис"'
            sheet['A4'] = "Борисюк К. М."

        if self.Company == 'ООО "Сисим"':
            sheet['A3'] = 'ООО "Сисим"'
            sheet['A4'] = "Ковалькову М. Н."

        if self.Company == 'ООО "СпецПодряд"':
            sheet['A3'] = 'ООО "СпецПодряд"'
            sheet['A4'] = "Осс А. В."

        if self.Company == 'ООО "КрасИнтегра"':
            sheet['A3'] = 'ООО "КрасИнтегра"'
            sheet['A4'] = "Сапину В. Д."

        # Сохранение файла
        workbook.save(file_path)

        # Печать файла
        printer = ExcelPrinter(self.server_address, self.smb_share)
        printer.print_excel_file(file_path)

    def print_doc_obhod(self, file_path):

        current_date = datetime.now().date()
        print_data = current_date.strftime('%d.%m.%Y')

        # Открытие файла с изменением атрибутов для доступа
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Формирование статичных данных
        sheet['E5'] = self.Dolznost
        sheet['E6'] = self.FIO
        sheet['B3'] = self.Uchastok
        sheet['J3'] = f"Дата {print_data} г."
        sheet['C6'] = self.state_obhod
        sheet['D12'] = self.Kladovchik
        sheet['D23'] = self.Kladovchik
        sheet['D33'] = self.SysAdmin
        sheet['D41'] = self.NachUch
        sheet['G36'] = f"{self.Dni} - дней"

        # Сохранение файла
        workbook.save(file_path)

        # Печать файла
        printer = ExcelPrinter(self.server_address, self.smb_share)
        printer.print_excel_file(file_path)

    def print_doc_uval(self, file_path):
        current_date = datetime.now().date()
        print_data = current_date.strftime('%d.%m.%Y')

        # Открытие файла с изменением атрибутов для доступа
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Формирование статичных данных
        sheet['A5'] = f"От {self.Dolznost}"
        sheet['A6'] = self.FIO
        sheet['A15'] = f"С {print_data}"

        # Проверка условия для Company
        if self.Company == 'ООО "А-Сервис"':
            sheet['A3'] = 'ООО "А-Сервис"'
            sheet['A4'] = "Белову В. Ю."

        if self.Company == 'ООО "Еда"':
            sheet['A3'] = 'ООО "Еда"'
            sheet['A4'] = "Потапову М. Г."

        if self.Company == 'ООО "Буфет"':
            sheet['A3'] = 'ООО "Буфет"'
            sheet['A4'] = "Потапову Г. В."

        if self.Company == 'ООО "ГМК Ангара"':
            sheet['A3'] = 'ООО "ГМК Ангара"'
            sheet['A4'] = "Калдарумиди Р. П."

        if self.Company == 'ООО "КомплектСервис"':
            sheet['A3'] = 'ООО "КомплектСервис"'
            sheet['A4'] = "Потриденный В. Ф."

        if self.Company == 'ООО "МоторСервис"':
            sheet['A3'] = 'ООО "МоторСервис"'
            sheet['A4'] = "Борисюк К. М."

        if self.Company == 'ООО "Сисим"':
            sheet['A3'] = 'ООО "Сисим"'
            sheet['A4'] = "Ковалькову М. Н."

        if self.Company == 'ООО "СпецПодряд"':
            sheet['A3'] = 'ООО "СпецПодряд"'
            sheet['A4'] = "Осс А. В."

        if self.Company == 'ООО "КрасИнтегра"':
            sheet['A3'] = 'ООО "КрасИнтегра"'
            sheet['A4'] = "Сапину В. Д."

        # Сохранение файла
        workbook.save(file_path)

        # Печать файла
        printer = ExcelPrinter(self.server_address, self.smb_share)
        printer.print_excel_file(file_path)

    def closeEvent(self, event):
        self.save_data_to_db()
        self.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
